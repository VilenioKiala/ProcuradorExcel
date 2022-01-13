from PyQt5.QtCore import QFile,Qt;
from PyQt5.uic import loadUi;
from PyQt5.QtWidgets import QApplication, QLabel,QMainWindow,QFileDialog,QMessageBox,QProgressDialog,QDialog,QTableWidget,QTableWidgetItem
import sys
from ProcuradorExcel_ui import Ui_MainWindow
import openpyxl;



class window(QMainWindow,Ui_MainWindow):
    def __init__(self,parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.connectButtonSignals()
        self.connectSignalsSlots()


    def connectSignalsSlots(self):
        self.actionAbout.triggered.connect(self.about)
        self.actionSair.triggered.connect(self.sair)


    def connectButtonSignals(self):
        self.btnAdicionarFicheiro.clicked.connect(self.openFileDialog);
        self.btnAdicionarPalavra.clicked.connect(self.adicionarPalavra);
        self.btnPesquisar.clicked.connect(self.pesquisarPalavras);

        self.btnRemoverFicheiro.clicked.connect(self.removerFicheiro);
        self.btnRemoverPalavra.clicked.connect(self.removerPalavra);

        self.btnAlterarPalavra.clicked.connect(self.alterarPalavra);


    def sair(self):
        self.close()

    def about(self):
        msgBox = QMessageBox(self)
        msgBox.setWindowFlag(Qt.FramelessWindowHint)
        msgBox.setText("meu texto qualquer vilenio boy")
        msgBox.setStyleSheet("padding: 12px; background-color: #201335;box-shadow: 1px 1px 1px black;")
        msgBox.exec()


    def openFileDialog(self):
        fileDialog = QFileDialog(parent=self)
        files,_ = fileDialog.getOpenFileNames(filter="*.xlsx")
        self.listFicheiros.addItems(files)

    def removerFicheiro(self):
        if self.listFicheiros.currentRow() >= 0:
            qntdItems = self.listFicheiros.count()
            items = [];
            for i in range(qntdItems):
                item = self.listFicheiros.item(i)
                items.append(item.text())

            del items[self.listFicheiros.currentRow()]
            self.listFicheiros.clear();
            self.listFicheiros.addItems(items)

    def adicionarPalavra(self):
        palavra = self.txtPalavra.text().strip()
        if palavra:
            self.listPalavras.addItem(palavra);
            self.txtPalavra.setText("")
        self.txtPalavra.setFocus()


    def removerPalavra(self):
        if self.listPalavras.currentRow() >= 0:
            qntdItems = self.listPalavras.count()
            items = [];
            for i in range(qntdItems):
                item = self.listPalavras.item(i)
                items.append(item.text())

            del items[self.listPalavras.currentRow()]
            self.listPalavras.clear();
            self.listPalavras.addItems(items)

    def alterarPalavra(self):
        if self.listPalavras.currentRow() >= 0:
            item = self.listPalavras.item(self.listPalavras.currentRow())
            item.setText(self.txtPalavra.text())


    def pesquisarPalavras(self):
        if self.listFicheiros.count() == 0:
            msgBox = QMessageBox(self)
            msgBox.setText("Nenhum ficheiro de pesquisa foi adicionado")
            msgBox.setWindowTitle("Aviso!")
            msgBox.show();
            return;

        if self.listPalavras.count() == 0:
            msgBox = QMessageBox(self)
            msgBox.setText("Coloque alguma palavra para pesquisar primeiro")
            msgBox.setWindowTitle("Aviso!")
            msgBox.show();
            return;

        progressDialog = QProgressDialog(self);
        progressDialog.setLabelText("asidsadas")
        progressDialog.exec()

        qntdFicheiros = self.listFicheiros.count()
        qntdPalavras = self.listPalavras.count()
        dadosPalavrasEncontradas = []

        for i in range(qntdFicheiros):
            ficheiro = self.listFicheiros.item(i).text()

            for j in range(qntdPalavras):
                palavra = self.listPalavras.item(j).text()
                wb = openpyxl.load_workbook(ficheiro)

                for ws in wb.worksheets:
                    for row in range(1,ws.max_row+1):
                        for col in range(1,ws.max_column+1):
                            valorCell = ws.cell(row=row,column=col).value
                            if  valorCell != None:
                                if valorCell.lower() == palavra.lower():
                                    dadosPalavrasEncontradas.append({
                                        "Palavra": palavra,
                                        "Ficheiro": ficheiro.split("/")[-1],
                                        "Linha": row,
                                        "Coluna":col,
                                    })

        progressDialog.close()

        if not dadosPalavrasEncontradas:
            msgBox2 = QMessageBox(self)
            msgBox2.setText("Nenhuma dessas palavras foram encontradas")
            msgBox2.setWindowTitle("Aviso!")
            msgBox2.exec();
            return;
        
        dialog = ResultsDialog(self)
        dialog.setWindowTitle("Resultados")
        dialog.tblResultados.setHorizontalHeaderLabels(["Palavra","Ficheiro","Linha","Coluna"])
        dialog.tblResultados.setRowCount(len(dadosPalavrasEncontradas))


        for i in range(len(dadosPalavrasEncontradas)):
            dialog.tblResultados.setItem(i,0,QTableWidgetItem(str(dadosPalavrasEncontradas[i]["Palavra"])))
            dialog.tblResultados.setItem(i,1,QTableWidgetItem(str(dadosPalavrasEncontradas[i]["Ficheiro"])))
            dialog.tblResultados.setItem(i,2,QTableWidgetItem(str(dadosPalavrasEncontradas[i]["Linha"])))
            dialog.tblResultados.setItem(i,3,QTableWidgetItem(str(dadosPalavrasEncontradas[i]["Coluna"])))

        
        dialog.exec()
    

class ResultsDialog(QDialog):
    def __init__(self,parent=None):
        super().__init__(parent);
        loadUi("ResultsDialog.ui",self);




app = QApplication(sys.argv)
window = window();
window.show()

app.exec()

